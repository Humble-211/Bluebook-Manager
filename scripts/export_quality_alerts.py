"""
Export Quality Alerts to Excel.

Scans all quality_alerts files in the database, filters to those following
the QA-YY-NNN naming convention, and exports to an Excel file.

Also updates the qa_counter table to reflect the highest QA number found.

Usage:
    python scripts/export_quality_alerts.py
"""

import os
import re
import sys
from datetime import datetime

# Ensure project root is importable and config.BASE_DIR resolves correctly.
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)
sys.argv[0] = os.path.join(PROJECT_ROOT, "main.py")

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from dal.database import init_db
from dal import dal

# QA filename pattern: QA-YY-NNN-DIENUM-description.docx
# Captures: year (YY), sequence (NNN), die number, and description
QA_PATTERN = re.compile(
    r"^QA-(\d{2})-(\d{3})-(\d+)-(.+)\.docx$",
    re.IGNORECASE,
)

SECTION_TYPE = "quality_alerts"


def parse_qa_filename(filename: str):
    """Parse a QA filename and return (year, seq_num, die_number, description) or None."""
    m = QA_PATTERN.match(filename)
    if not m:
        return None
    year = int(m.group(1))
    seq_num = int(m.group(2))
    die_number = m.group(3)
    description = m.group(4).replace("-", " ")
    return year, seq_num, die_number, description


def main():
    init_db()

    # Get all bluebooks
    all_bluebooks = dal.list_bluebooks(limit=10000)
    print(f"Found {len(all_bluebooks)} bluebook(s) in database.\n")

    # Collect QA data
    qa_rows = []
    max_seq_by_year = {}  # track highest QA number per year
    skipped = 0

    for bb in all_bluebooks:
        files = dal.get_files_for_bluebook(bb.id, SECTION_TYPE)
        for bf in files:
            filename = os.path.basename(bf.file_path)
            parsed = parse_qa_filename(filename)
            if not parsed:
                skipped += 1
                continue

            year, seq_num, die_number, description = parsed

            # Track the highest sequence number per year
            if year not in max_seq_by_year or seq_num > max_seq_by_year[year]:
                max_seq_by_year[year] = seq_num

            # Get customer names from the bluebook
            customers = ", ".join(bb.customer_names) if bb.customer_names else ""

            # Get file date from created_at
            file_date = bf.created_at if hasattr(bf, "created_at") and bf.created_at else ""

            qa_rows.append({
                "filename": filename,
                "die": die_number,
                "customer": customers,
                "description": description,
                "date": file_date,
                "year": year,
                "seq": seq_num,
            })

    # Sort by year descending, then sequence descending
    qa_rows.sort(key=lambda r: (r["year"], r["seq"]), reverse=True)

    # Update qa_counter in database for each year found
    for year, max_num in max_seq_by_year.items():
        current = dal.get_qa_counter(year)
        if max_num > current:
            dal.set_qa_counter(year, max_num)
            print(f"  Updated QA counter: year 20{year:02d} → {max_num}")
        else:
            print(f"  QA counter for 20{year:02d}: {current} (unchanged)")

    if not qa_rows:
        print("\nNo quality alerts with QA naming convention found.")
        return

    # ── Create Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Quality Alerts"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Filename", "Die", "Customer", "Description", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(vertical="center", wrap_text=True)

    for row_idx, qa in enumerate(qa_rows, 2):
        values = [
            qa["filename"],
            qa["die"],
            qa["customer"],
            qa["description"],
            qa["date"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 50  # Filename
    ws.column_dimensions["B"].width = 12  # Die
    ws.column_dimensions["C"].width = 25  # Customer
    ws.column_dimensions["D"].width = 50  # Description
    ws.column_dimensions["E"].width = 18  # Date

    # Save
    output_path = os.path.join(PROJECT_ROOT, "quality_alerts_export.xlsx")
    wb.save(output_path)

    print(f"\n{'='*60}")
    print(f"  Exported: {len(qa_rows)} quality alert(s)")
    print(f"  Skipped:  {skipped} file(s) (non-QA naming)")
    print(f"  File:     {output_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
